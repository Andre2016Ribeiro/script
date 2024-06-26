import mechanicalsoup

 
from openpyxl import load_workbook


data_file = 'artigo.xlsx'

# Load the entire workbook.
wb = load_workbook(data_file)

# Load one worksheet.
ws = wb.worksheets[2]

all_rows = list(ws.rows)
row1 = 3  
 
for row in all_rows[2:7973]:
    print(row[0].value)


    browser = mechanicalsoup.Browser()
    url = "http://10.0.0.152:88//GPR/ProcDesen.php/?artigo="+row[0].value
    page = browser.get(url)
    
   
    try: 
        tag = page.soup.select("Input")[0]["value"]
        mycell1= ws.cell(row=row1, column=9)  
        mycell1.value=len(tag)
        print(tag)
         
        mycell= ws.cell(row=row1, column=10)  
        mycell.value=tag
       
        mycell2= ws.cell(row=row1, column=11)  
        mycell2.value=url
        mycell2.hyperlink=url
    except:
        mycell2= ws.cell(row=row1, column=11)  
        mycell2.value=url
        mycell2.hyperlink=url
        mycell1= ws.cell(row=row1, column=9)  
        mycell1.value='sucesso'
        pass
    
         

    row1+=1
   
    
    

   
wb.save(data_file)



